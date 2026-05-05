"""One-shot import: pull 5-month daily_summary history from the source xlsm.

Reads sheet 「綜合整理」 — already-aggregated row-per-day data, exactly
matching the daily_summary schema after a column rename.
"""
import sys, pathlib, argparse
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent.parent))
from openpyxl import load_workbook
from app.db import connect, init_db, DB_PATH


def parse_md_label(label: str, fallback_year: int = 2025) -> str | None:
    """'11_7' / '12_1' / '1_5' / '4_16' → 'YYYY-MM-DD' (heuristic year boundary)."""
    if not label or "_" not in str(label):
        return None
    parts = str(label).split("_")
    if len(parts) != 2:
        return None
    try:
        m, d = int(parts[0]), int(parts[1])
    except ValueError:
        return None
    # The Excel covers 11/7 ~ 4/16. Months 11–12 belong to year base, 1–4 to base+1.
    year = fallback_year if m >= 7 else fallback_year + 1
    return f"{year:04d}-{m:02d}-{d:02d}"


def main(xlsm_path: str, base_year: int) -> None:
    init_db()
    wb = load_workbook(xlsm_path, data_only=True, read_only=True, keep_vba=False)
    ws = wb["綜合整理"]
    rows = list(ws.iter_rows(values_only=True))

    # Header layout (R1+R2 merged):
    #   col 0  '' (label)            -> date
    #   col 1  '前一日日盤Data'        -> ignore (just prev label)
    #   col 2  '日盤收盤'             -> tx_close
    #   col 3  '法人淨部位'           -> op_legal_net
    #   col 4  '法人買權CALL'         -> op_call_net
    #   col 5  '法人賣權PUT'          -> op_put_net
    #   col 6  '法人CP合計多空'       -> op_cp_net
    #   col 7  '開盤前多空'           -> fut_pre_open_net
    #   col 8  '股票期貨 法人淨部位'   -> stock_fut_legal_net
    #   col 9  '融資佔市值比 上市'    -> twse_margin_pct
    #   col 10 '融資佔市值比 上櫃'    -> tpex_margin_pct
    #   col 11 '融資餘額 上市(億元)'  -> twse_margin_amt_oku
    #   col 12 '融資餘額 上櫃(億元)'  -> tpex_margin_amt_oku
    #   col 13 '總市值 上市(兆元)'    -> twse_mkt_cap_chao
    #   col 14 '總市值 上櫃(兆元)'    -> tpex_mkt_cap_chao
    inserted = 0
    skipped = 0
    with connect() as con:
        for row in rows[2:]:  # skip 2 header rows
            # Col 0 = 看盤日期 (e.g. '12_1'), Col 1 = 資料日期 (e.g. '11_28').
            # Use 資料日期 — that's the trading date the numbers describe.
            label = row[1] or row[0]
            if not label:
                continue
            iso_date = parse_md_label(label, fallback_year=base_year)
            if not iso_date:
                skipped += 1
                continue
            con.execute("""
                INSERT OR REPLACE INTO daily_summary
                (date, tx_close, op_legal_net, op_call_net, op_put_net, op_cp_net,
                 fut_pre_open_net, stock_fut_legal_net,
                 twse_margin_pct, tpex_margin_pct,
                 twse_margin_amt_oku, tpex_margin_amt_oku,
                 twse_mkt_cap_chao, tpex_mkt_cap_chao)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                iso_date,
                row[2], row[3], row[4], row[5], row[6],
                row[7], row[8],
                row[9], row[10],
                row[11], row[12],
                row[13], row[14],
            ))
            inserted += 1
    print(f"Imported {inserted} rows into daily_summary  (skipped {skipped}).")
    print(f"DB: {DB_PATH}")


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--xlsm", default=r"D:/ClaudeCode/法人OP日夜盤數據/_analysis/source.xlsm")
    p.add_argument("--base-year", type=int, default=2025,
                   help="Year for months 7-12 (others go to base_year+1)")
    args = p.parse_args()
    main(args.xlsm, args.base_year)
